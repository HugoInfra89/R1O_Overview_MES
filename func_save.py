def save_data_to_local(dataframe, path=r"C:\Users\hvanegeraat\AppData\Local\Temp"):
        folderpath_to_temp_folder = path
        file_name = "Overview MES.xlsx"

        path_temp_excel_file = rf"{folderpath_to_temp_folder}\{file_name}"
        dataframe.to_excel(path_temp_excel_file, sheet_name="Overview MES", index=False)

def open_from_local_and_modify(file_name="Overview MES.xlsx",
        project_data = "",
        path_local=r"C:\Users\hvanegeraat\AppData\Local\Temp",
        path_sharepoint=r"C:\Users\hvanegeraat\THV ROCO\R1O (050) - 04 BIM\Overview objects ES R1O\All MES overview R1O"):

        from openpyxl import load_workbook
        path_to_temp_file = rf"{path_local}\{file_name}"
        print(path_to_temp_file)
        wb = load_workbook(path_to_temp_file)
        ws = wb.active

        # Change the width of the columns
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 50

        modelname_columns = ["D","F","H","J","L","N"]
        for model_column_name in modelname_columns:
                ws.column_dimensions[model_column_name].width = 40

        bool_columns = ["E", "G", "I", "K", "M", "O"]
        for bool_col_name in bool_columns:
                ws.column_dimensions[bool_col_name].width = 20

        folderpath_sharepoint = path_sharepoint
        path_sharepoint = rf"{folderpath_sharepoint}/{file_name}"

        def make_metadata_dict(project_data=("XXX", "(000)_XXX")):
                import datetime as dt
                current_datetime = dt.datetime.now()
                formatted_time = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
                project_area = project_data[0]
                project_code = project_data[-1]

                meta_data_dict = {
                        "Project Code" : project_code,
                        "Project Area" : project_area,
                        "Date Time" : formatted_time.split()[0],
                        "Time": formatted_time.split()[-1]}
                return meta_data_dict

        md_dict = make_metadata_dict()

        ws_md = wb.create_sheet(title="Metadata")
        wb.active = ws_md

        print("Active Worksheet:", ws_md.title)

        ws_md.append(list(md_dict.keys()))
        wb.save(path_sharepoint)


        import os
        os.system(f'start excel "{path_sharepoint}"')

def save_bar_chart_on_sharepoint(plt):
        path = r'C:\Users\hvanegeraat\THV ROCO\R1O (050) - 04 BIM\Overview objects ES R1O\All MES overview R1O'
        file_name = "R1O_barchart.jpg"
        file_path = path + file_name

        print(file_path)
        plt.show()
        breakpoint()

        plt.savefig(file_path)

